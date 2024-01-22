from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from openpyxl import Workbook, load_workbook
import time
import pyperclip
nav = webdriver.Chrome()
nav.get("https://www.kalunga.com.br/busca/1?q=teclado")
time.sleep(3)

# //tag[@atributo='valor']

titulos = nav.find_elements(By.XPATH, "//h2[@class='blocoproduto__title mb-0 mt-2 pb-2 pb-lg-3']")
precos = nav.find_elements(By.XPATH, "//span[@class='blocoproduto__text blocoproduto__text--bold blocoproduto__price']")
link = nav.find_elements(By.XPATH, "//a[@class='blocoproduto__link']")

workbook = Workbook()
workbook.create_sheet('produtos')
sheetProdutos = workbook['produtos']
sheetProdutos['A1'].value = 'Nome'
sheetProdutos['B1'].value = 'Preço'
sheetProdutos['C1'].value = 'Link'

for titulo, preco, site in zip(titulos, precos, link):
    sheetProdutos.append([titulo.text, preco.text, site.get_attribute('href')])

workbook.save('produtos.xlsx')
time.sleep(3)

nav.get("https://web.whatsapp.com")
time.sleep(45)

nav.find_element('xpath', '/html/body/div[1]/div/div[2]/div[3]/div/div[1]/div/div[2]/div[2]/div').click()
nav.find_element('xpath', '//*[@id="side"]/div[1]/div/div[2]/div[2]/div/div[1]/p').send_keys("Solo")
nav.find_element('xpath', '//*[@id="side"]/div[1]/div/div[2]/div[2]/div/div[1]/p').send_keys(Keys.ENTER)
time.sleep(10)

planilha = load_workbook('produtos.xlsx')
sheet = planilha['produtos']

for nome, preco, link in zip(sheet["A"], sheet["B"], sheet["C"]):
    pyperclip.copy(nome.value + " " + preco.value + "  " + link.value)
    time.sleep(3)
    nav.find_element('xpath', '//*[@id="main"]/footer/div[1]/div/span[2]/div/div[2]/div[1]/div[2]/div[1]/p').send_keys(Keys.CONTROL + "v")
    time.sleep(2)
    nav.find_element('xpath', '//*[@id="main"]/footer/div[1]/div/span[2]/div/div[2]/div[1]/div[2]/div[1]/p').send_keys(Keys.ENTER)
    time.sleep(5)